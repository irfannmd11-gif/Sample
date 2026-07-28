"""Download, inspect, redact, and format Excel workbooks for the UI."""

from __future__ import annotations

import hashlib
import re
from pathlib import Path
from urllib.parse import urlparse
from urllib.request import Request, urlopen

from openpyxl import load_workbook

from config import Settings


class WorkbookError(RuntimeError):
    pass


class WorkbookDownloader:
    def __init__(self, settings: Settings) -> None:
        self.settings = settings

    def download(self, url: str, project_reference: str, task_number: str) -> Path:
        parsed = urlparse(url)
        if parsed.scheme not in {"http", "https"}:
            raise WorkbookError("The workbook link must use HTTP or HTTPS.")
        if not parsed.netloc:
            raise WorkbookError("The workbook link is invalid.")

        self.settings.download_dir.mkdir(parents=True, exist_ok=True)
        token = hashlib.sha256(f"{project_reference}:{task_number}".encode()).hexdigest()[:16]
        destination = self.settings.download_dir / f"{token}.xlsx"
        temporary = destination.with_suffix(".part")
        request = Request(url, headers={"User-Agent": "WorkbookValidationTool/1.0"})

        try:
            with urlopen(request, timeout=30) as response, temporary.open("wb") as output:
                content_length = response.headers.get("Content-Length")
                if content_length and int(content_length) > self.settings.max_workbook_bytes:
                    raise WorkbookError("The workbook exceeds the configured size limit.")
                copied = 0
                while chunk := response.read(1024 * 1024):
                    copied += len(chunk)
                    if copied > self.settings.max_workbook_bytes:
                        raise WorkbookError("The workbook exceeds the configured size limit.")
                    output.write(chunk)
        except WorkbookError:
            temporary.unlink(missing_ok=True)
            raise
        except OSError as error:
            temporary.unlink(missing_ok=True)
            raise WorkbookError("The workbook could not be downloaded.") from error

        temporary.replace(destination)
        return destination


class WorkbookProcessor:
    MAX_SAMPLE_ROWS = 5
    MAX_SAMPLE_COLUMNS = 20
    PII_PATTERNS = (
        ("email", re.compile(r"[A-Z0-9._%+-]+@[A-Z0-9.-]+\.[A-Z]{2,}", re.I)),
        ("phone", re.compile(r"(?<!\d)(?:\+?\d{1,3}[ -]?)?\d{10}(?!\d)")),
        ("aadhaar", re.compile(r"(?<!\d)\d{4}[ -]?\d{4}[ -]?\d{4}(?!\d)")),
        ("card", re.compile(r"(?<!\d)(?:\d[ -]?){13,16}(?!\d)")),
    )

    def build_payload(self, workbook_path: Path, offer_description: str = "") -> dict:
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=False)
        except Exception as error:
            raise WorkbookError("The downloaded file is not a readable Excel workbook.") from error

        pii_count = 0
        sheet_data: dict[str, dict] = {}
        try:
            for sheet in workbook.worksheets:
                rendered, found = self._sheet_sample(sheet)
                pii_count += found
                dimensions = f"{sheet.max_row} rows × {sheet.max_column} columns"
                sheet_data[sheet.title] = {
                    "offer_description": offer_description,
                    "workbook_context": self._workbook_context(workbook_path.name, workbook.sheetnames),
                    "sheet_prompt": self._sheet_prompt(sheet.title, dimensions, rendered),
                    "generated_prompt": self._generated_prompt(offer_description, workbook_path.name, sheet.title, dimensions, rendered),
                }
        finally:
            workbook.close()

        return {
            "workbook_name": workbook_path.name,
            "workbook_type": workbook_path.suffix.removeprefix(".").upper() or "XLSX",
            "sheet_count": len(workbook.sheetnames),
            "sheets": workbook.sheetnames,
            "sheet_data": sheet_data,
            "pii_status": f"{pii_count} masked" if pii_count else "None detected",
        }

    def _sheet_sample(self, sheet) -> tuple[str, int]:
        rows: list[str] = []
        pii_count = 0
        for row in sheet.iter_rows(max_row=self.MAX_SAMPLE_ROWS, max_col=self.MAX_SAMPLE_COLUMNS, values_only=True):
            values = []
            for value in row:
                masked, found = self._mask_pii("" if value is None else str(value))
                pii_count += found
                values.append(masked)
            rows.append(" | ".join(values).rstrip(" |"))
        return "\n".join(rows).strip() or "(The sheet is empty.)", pii_count

    def _mask_pii(self, text: str) -> tuple[str, int]:
        matches = 0
        for _, pattern in self.PII_PATTERNS:
            text, count = pattern.subn("[REDACTED]", text)
            matches += count
        return text, matches

    @staticmethod
    def _workbook_context(name: str, sheets: list[str]) -> str:
        return f"Workbook: {name}\nAvailable sheets: {', '.join(sheets)}\nPII in the displayed sample has been masked."

    @staticmethod
    def _sheet_prompt(name: str, dimensions: str, sample: str) -> str:
        return f"Validate the '{name}' sheet ({dimensions}). Review structure, required fields, data types, formulas, and values.\n\nMasked sample:\n{sample}"

    @staticmethod
    def _generated_prompt(offer: str, name: str, sheet: str, dimensions: str, sample: str) -> str:
        return (
            "# Workbook Validation Prompt\n\n"
            f"## Offer Description\n{offer or '(Not available)'}\n\n"
            f"## Workbook Context\nWorkbook: {name}\nSheet: {sheet} ({dimensions})\n\n"
            f"## Masked Sheet Sample\n{sample}\n\n"
            "## Expected Output\nList validation issues with row/column references, severity, and suggested remediation."
        )
